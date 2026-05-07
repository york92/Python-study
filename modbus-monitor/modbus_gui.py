"""
Modbus 上位机 GUI
- 启动时读取 device-config.xlsx，按 Sheet → Tab 渲染页面
- 串口配置区（顶部常驻）
- 每个 Tab 是 N行×4列 的表单网格
- Modbus 读写接口已预留（mock 实现，替换 ModbusClient 即可）
"""

import sys
import json
import math
import os
import random
from typing import Any, Dict, List, Optional

import openpyxl
from PyQt5.QtCore import (Qt, QTimer, QThread, pyqtSignal, QObject,
                           QPropertyAnimation, QEasingCurve)
from PyQt5.QtGui import (QFont, QColor, QPalette, QIcon, QPainter,
                         QBrush, QPen, QPixmap, QFontDatabase)
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QLineEdit, QComboBox, QFrame, QTabWidget,
    QScrollArea, QGroupBox, QSizePolicy, QSpacerItem, QStatusBar,
    QMessageBox, QPushButton, QSpinBox
)

# ─────────────────────────────────────────────────────────────────
#  样式表 (深色工业风 UI)
# ─────────────────────────────────────────────────────────────────
STYLE = """
QMainWindow, QWidget {
    background-color: #1a1d23;
    color: #e0e4ef;
    font-family: 'Microsoft YaHei', 'PingFang SC', 'Noto Sans CJK SC', sans-serif;
    font-size: 13px;
}

/* ── 顶部串口配置栏 ── */
#serialBar {
    background-color: #22262f;
    border-bottom: 2px solid #2e7d9e;
    padding: 6px 12px;
}
#serialBar QLabel {
    color: #8eb4cc;
    font-size: 12px;
    font-weight: 600;
    letter-spacing: 0.5px;
}
#serialBar QComboBox, #serialBar QSpinBox {
    background-color: #2c3140;
    color: #d0dff0;
    border: 1px solid #3a4560;
    border-radius: 4px;
    padding: 3px 8px;
    min-width: 90px;
    selection-background-color: #2e7d9e;
}
#serialBar QComboBox:hover, #serialBar QSpinBox:hover {
    border-color: #2e7d9e;
}
#serialBar QComboBox::drop-down {
    border: none;
    width: 20px;
}
#serialBar QComboBox QAbstractItemView {
    background-color: #2c3140;
    border: 1px solid #2e7d9e;
    selection-background-color: #2e7d9e;
    color: #d0dff0;
}

/* 连接按钮 */
#btnConnect {
    background-color: #1d6e42;
    color: #a8f0c6;
    border: 1px solid #27a561;
    border-radius: 5px;
    padding: 5px 18px;
    font-weight: 700;
    font-size: 13px;
    min-width: 80px;
}
#btnConnect:hover { background-color: #27a561; }
#btnConnect:pressed { background-color: #156035; }
#btnConnect[connected="true"] {
    background-color: #7a1c1c;
    color: #f0a8a8;
    border-color: #c13030;
}
#btnConnect[connected="true"]:hover { background-color: #c13030; }

/* 状态指示灯 */
#statusLed {
    min-width: 14px; max-width: 14px;
    min-height: 14px; max-height: 14px;
    border-radius: 7px;
    background-color: #3a3a3a;
    border: 1px solid #555;
}
#statusLed[connected="true"] {
    background-color: #27c96a;
    border-color: #1aff7a;
}

/* ── Tab ── */
QTabWidget::pane {
    border: none;
    background-color: #1a1d23;
}
QTabBar::tab {
    background-color: #22262f;
    color: #7a8aab;
    padding: 8px 22px;
    margin-right: 2px;
    border-top-left-radius: 5px;
    border-top-right-radius: 5px;
    border: 1px solid #2e333f;
    border-bottom: none;
    font-weight: 600;
    font-size: 13px;
}
QTabBar::tab:selected {
    background-color: #1a1d23;
    color: #4fc3f7;
    border-color: #2e7d9e;
    border-bottom: 2px solid #1a1d23;
}
QTabBar::tab:hover:!selected { background-color: #2a2f3d; color: #a0b4cc; }

/* ── 滚动区 ── */
QScrollArea { border: none; background-color: transparent; }
QScrollBar:vertical {
    background-color: #22262f;
    width: 8px; border-radius: 4px;
}
QScrollBar::handle:vertical {
    background-color: #3a4560;
    border-radius: 4px; min-height: 30px;
}
QScrollBar::handle:vertical:hover { background-color: #2e7d9e; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }

/* ── 表单单元格 ── */
#cell {
    background-color: #22262f;
    border: 1px solid #2e333f;
    border-radius: 6px;
    margin: 4px;
}
#cell:hover { border-color: #3a5070; }

#cellLabel {
    color: #8499bb;
    font-size: 12px;
    padding: 0 0 2px 0;
}
#cellUnit {
    color: #4fc3f7;
    font-size: 11px;
    font-weight: 700;
}

/* 只读输入框 */
#inputReadonly {
    background-color: #181b21;
    color: #4fc3f7;
    border: 1px solid #2a3040;
    border-radius: 4px;
    padding: 4px 8px;
    font-size: 14px;
    font-weight: 600;
    font-family: 'Courier New', monospace;
}

/* 可写输入框 */
#inputWrite {
    background-color: #1e2535;
    color: #ffd580;
    border: 1px solid #3a4d6a;
    border-radius: 4px;
    padding: 4px 8px;
    font-size: 14px;
    font-weight: 600;
}
#inputWrite:hover { border-color: #4fc3f7; }
#inputWrite:focus {
    border-color: #ffd580;
    background-color: #252d40;
}

/* 下拉选 */
#selectWrite {
    background-color: #1e2535;
    color: #ffd580;
    border: 1px solid #3a4d6a;
    border-radius: 4px;
    padding: 4px 8px;
    font-size: 13px;
    font-weight: 600;
}
#selectWrite:hover { border-color: #4fc3f7; }
#selectWrite:focus { border-color: #ffd580; }
#selectWrite::drop-down { border: none; width: 22px; }
#selectWrite QAbstractItemView {
    background-color: #1e2535;
    border: 1px solid #4fc3f7;
    selection-background-color: #2e7d9e;
    color: #ffd580;
}

/* ── 状态栏 ── */
QStatusBar {
    background-color: #161920;
    color: #556688;
    border-top: 1px solid #2e333f;
    font-size: 12px;
    padding: 2px 8px;
}
"""

# ─────────────────────────────────────────────────────────────────
#  数据模型
# ─────────────────────────────────────────────────────────────────
class ChannelConfig:
    """一条配置条目"""
    def __init__(self, name, register, decimals, unit, access, widget_type, options):
        self.name: str = str(name).strip()
        self.register: int = int(register)
        self.decimals: int = int(decimals)
        self.unit: str = str(unit).strip() if unit else ""
        self.access: str = str(access).strip().upper()   # READONLY / READWRITE
        self.widget_type: str = str(widget_type).strip().upper()  # INPUT / SELECT
        self.options: List[Dict] = []
        if options and str(options).strip():
            try:
                self.options = json.loads(str(options))
            except Exception:
                pass

    @property
    def readonly(self):
        return self.access == "READONLY"

    def format_value(self, raw: int) -> str:
        """寄存器原始值 → 显示字符串"""
        if self.decimals > 0:
            return f"{raw / (10 ** self.decimals):.{self.decimals}f}"
        return str(raw)

    def parse_value(self, text: str) -> Optional[int]:
        """显示字符串 → 寄存器原始值"""
        try:
            return int(round(float(text) * (10 ** self.decimals)))
        except Exception:
            return None


def load_config(path: str) -> Dict[str, List[ChannelConfig]]:
    """读取 xlsx，返回 {sheet_name: [ChannelConfig, ...]}"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        configs = []
        rows = list(ws.iter_rows(values_only=True))
        # 跳过表头行（判断第一行第一列是否为 'name'）
        start = 1 if (rows and str(rows[0][0]).strip().lower() == "name") else 0
        for row in rows[start:]:
            if not row or not row[0]:
                continue
            name, reg, dec, unit, access, wtype = (row[i] if i < len(row) else "" for i in range(6))
            opts = row[6] if len(row) > 6 else ""
            try:
                configs.append(ChannelConfig(name, reg, dec, unit, access, wtype, opts))
            except Exception as e:
                print(f"[WARN] 跳过无效行 {row}: {e}")
        if configs:
            result[sheet_name] = configs
    return result


# ─────────────────────────────────────────────────────────────────
#  Modbus 客户端（预留接口，当前为 Mock）
# ─────────────────────────────────────────────────────────────────
class ModbusClient:
    """
    预留接口类 —— 替换此类中的实现即可接入真实 Modbus RTU。
    推荐使用 pymodbus: from pymodbus.client import ModbusSerialClient
    """
    def __init__(self):
        self._connected = False
        self._mock_data: Dict[int, int] = {}

    def connect(self, port: str, baudrate: int, stopbits: int,
                bytesize: int, parity: str, slave_id: int = 1) -> bool:
        """
        TODO: 替换为真实串口连接
        client = ModbusSerialClient(
            port=port, baudrate=baudrate, stopbits=stopbits,
            bytesize=bytesize, parity=parity, timeout=1)
        self._connected = client.connect()
        """
        print(f"[MOCK] connect {port} @ {baudrate} baud")
        self._connected = True
        return True

    def disconnect(self):
        print("[MOCK] disconnect")
        self._connected = False

    @property
    def is_connected(self):
        return self._connected

    def read_register(self, address: int, slave_id: int = 1) -> Optional[int]:
        """
        TODO: 替换为真实读取
        result = client.read_holding_registers(address - 40001, 1, slave=slave_id)
        if not result.isError(): return result.registers[0]
        """
        if not self._connected:
            return None
        # Mock: 随机抖动
        base = self._mock_data.get(address, random.randint(100, 9999))
        val = max(0, base + random.randint(-5, 5))
        self._mock_data[address] = val
        return val

    def write_register(self, address: int, value: int, slave_id: int = 1) -> bool:
        """
        TODO: 替换为真实写入
        result = client.write_register(address - 40001, value, slave=slave_id)
        return not result.isError()
        """
        if not self._connected:
            return False
        print(f"[MOCK] write {address} = {value}")
        self._mock_data[address] = value
        return True


# ─────────────────────────────────────────────────────────────────
#  轮询工作线程
# ─────────────────────────────────────────────────────────────────
class PollingWorker(QObject):
    values_ready = pyqtSignal(dict)   # {register: raw_int}
    finished = pyqtSignal()

    def __init__(self, client: ModbusClient,
                 registers: List[int], interval_ms: int):
        super().__init__()
        self.client = client
        self.registers = registers
        self.interval_ms = interval_ms
        self._running = False

    def run(self):
        self._running = True
        while self._running:
            data = {}
            for reg in self.registers:
                v = self.client.read_register(reg)
                if v is not None:
                    data[reg] = v
            self.values_ready.emit(data)
            if self.interval_ms > 0:
                ms = self.interval_ms
                while ms > 0 and self._running:
                    QThread.msleep(min(ms, 50))
                    ms -= 50
        self.finished.emit()

    def stop(self):
        self._running = False


# ─────────────────────────────────────────────────────────────────
#  单元格 Widget
# ─────────────────────────────────────────────────────────────────
class CellWidget(QFrame):
    write_requested = pyqtSignal(int, int)  # (register, raw_value)

    def __init__(self, cfg: ChannelConfig, parent=None):
        super().__init__(parent)
        self.cfg = cfg
        self.setObjectName("cell")
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 8, 10, 8)
        layout.setSpacing(4)

        # ── 标题行 ──
        header = QHBoxLayout()
        header.setSpacing(4)

        lbl = QLabel(self.cfg.name)
        lbl.setObjectName("cellLabel")
        lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        header.addWidget(lbl)

        if self.cfg.unit:
            unit_lbl = QLabel(self.cfg.unit)
            unit_lbl.setObjectName("cellUnit")
            header.addWidget(unit_lbl)

        # 权限标记
        badge = QLabel("RO" if self.cfg.readonly else "RW")
        badge.setFixedSize(26, 16)
        badge.setAlignment(Qt.AlignCenter)
        badge.setStyleSheet(
            "background:#1a3a2a;color:#27c96a;border-radius:3px;font-size:10px;font-weight:700;"
            if not self.cfg.readonly else
            "background:#1a2a3a;color:#4a8ab0;border-radius:3px;font-size:10px;font-weight:700;"
        )
        header.addWidget(badge)
        layout.addLayout(header)

        # ── 输入控件 ──
        if self.cfg.widget_type == "SELECT" and self.cfg.options:
            self._widget = QComboBox()
            self._widget.setObjectName("selectWrite")
            for opt in self.cfg.options:
                self._widget.addItem(opt.get("label", ""), opt.get("value", ""))
            self._widget.currentIndexChanged.connect(self._on_select_changed)
            if self.cfg.readonly:
                self._widget.setEnabled(False)
        else:
            self._widget = QLineEdit()
            self._widget.setPlaceholderText("--")
            if self.cfg.readonly:
                self._widget.setObjectName("inputReadonly")
                self._widget.setReadOnly(True)
            else:
                self._widget.setObjectName("inputWrite")
                self._widget.returnPressed.connect(self._on_input_enter)

        self._widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self._widget.setFixedHeight(32)
        layout.addWidget(self._widget)

        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setMinimumHeight(80)

    def update_value(self, raw: int):
        """轮询到新值 → 刷新显示"""
        display = self.cfg.format_value(raw)
        if isinstance(self._widget, QComboBox):
            # SELECT: 匹配 value 字段
            raw_str = str(raw)
            for i in range(self._widget.count()):
                if str(self._widget.itemData(i)) == raw_str:
                    # 临时屏蔽信号，避免触发写操作
                    self._widget.blockSignals(True)
                    self._widget.setCurrentIndex(i)
                    self._widget.blockSignals(False)
                    break
        else:
            self._widget.setText(display)

    def _on_input_enter(self):
        text = self._widget.text().strip()
        raw = self.cfg.parse_value(text)
        if raw is not None:
            self.write_requested.emit(self.cfg.register, raw)
        else:
            self._widget.setStyleSheet("border-color: #c13030;")
            QTimer.singleShot(1000, lambda: self._widget.setStyleSheet(""))

    def _on_select_changed(self, idx: int):
        if self.cfg.readonly:
            return
        val_str = self._widget.itemData(idx)
        try:
            self.write_requested.emit(self.cfg.register, int(val_str))
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────
#  Tab 页：一个 Sheet 对应的表单
# ─────────────────────────────────────────────────────────────────
COLS = 4  # 每行 4 个单元格

class SheetTab(QScrollArea):
    write_requested = pyqtSignal(int, int)

    def __init__(self, configs: List[ChannelConfig], parent=None):
        super().__init__(parent)
        self.configs = configs
        self.cells: Dict[int, CellWidget] = {}  # register → CellWidget
        self.setWidgetResizable(True)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._build_ui()

    def _build_ui(self):
        container = QWidget()
        grid = QGridLayout(container)
        grid.setSpacing(6)
        grid.setContentsMargins(12, 12, 12, 12)

        rows = math.ceil(len(self.configs) / COLS)
        for idx, cfg in enumerate(self.configs):
            r, c = divmod(idx, COLS)
            cell = CellWidget(cfg)
            cell.write_requested.connect(self.write_requested)
            grid.addWidget(cell, r, c)
            self.cells[cfg.register] = cell

        # 填充尾部空格（保持对齐）
        total = rows * COLS
        for i in range(len(self.configs), total):
            r, c = divmod(i, COLS)
            spacer = QWidget()
            spacer.setObjectName("cell")
            spacer.setVisible(False)
            grid.addWidget(spacer, r, c)

        for c in range(COLS):
            grid.setColumnStretch(c, 1)

        self.setWidget(container)

    def apply_values(self, data: Dict[int, int]):
        for reg, raw in data.items():
            if reg in self.cells:
                self.cells[reg].update_value(raw)


# ─────────────────────────────────────────────────────────────────
#  串口配置栏
# ─────────────────────────────────────────────────────────────────
class SerialBar(QWidget):
    connect_clicked = pyqtSignal(dict)   # 参数字典
    disconnect_clicked = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("serialBar")
        self._connected = False
        self._build_ui()

    def _build_ui(self):
        h = QHBoxLayout(self)
        h.setContentsMargins(12, 6, 12, 6)
        h.setSpacing(10)

        def lbl(text):
            l = QLabel(text)
            l.setObjectName("serialBar")
            return l

        # 串口
        h.addWidget(lbl("串口:"))
        self.cb_port = QComboBox()
        self.cb_port.setObjectName("serialBar")
        self.cb_port.addItems(["COM1", "COM2", "COM3", "COM4",
                                "/dev/ttyUSB0", "/dev/ttyUSB1",
                                "/dev/ttyS0"])
        self.cb_port.setEditable(True)
        self.cb_port.setMinimumWidth(110)
        h.addWidget(self.cb_port)

        # 波特率
        h.addWidget(lbl("波特率:"))
        self.cb_baud = QComboBox()
        self.cb_baud.setObjectName("serialBar")
        self.cb_baud.addItems(["9600", "19200", "38400", "57600", "115200"])
        self.cb_baud.setCurrentText("9600")
        h.addWidget(self.cb_baud)

        # 数据位
        h.addWidget(lbl("数据位:"))
        self.cb_data = QComboBox()
        self.cb_data.setObjectName("serialBar")
        self.cb_data.addItems(["8", "7"])
        h.addWidget(self.cb_data)

        # 停止位
        h.addWidget(lbl("停止位:"))
        self.cb_stop = QComboBox()
        self.cb_stop.setObjectName("serialBar")
        self.cb_stop.addItems(["1", "2"])
        h.addWidget(self.cb_stop)

        # 校验
        h.addWidget(lbl("校验:"))
        self.cb_parity = QComboBox()
        self.cb_parity.setObjectName("serialBar")
        self.cb_parity.addItems(["None", "Odd", "Even"])
        h.addWidget(self.cb_parity)

        # 从站 ID
        h.addWidget(lbl("从站ID:"))
        self.sp_slave = QSpinBox()
        self.sp_slave.setObjectName("serialBar")
        self.sp_slave.setRange(1, 247)
        self.sp_slave.setValue(1)
        self.sp_slave.setFixedWidth(60)
        h.addWidget(self.sp_slave)

        # 轮询间隔
        h.addWidget(lbl("轮询(ms):"))
        self.sp_poll = QSpinBox()
        self.sp_poll.setObjectName("serialBar")
        self.sp_poll.setRange(0, 9999)
        self.sp_poll.setValue(500)
        self.sp_poll.setFixedWidth(70)
        self.sp_poll.setToolTip("0 = 无等待连续轮询")
        h.addWidget(self.sp_poll)

        h.addStretch()

        # 状态指示灯
        self.led = QLabel()
        self.led.setObjectName("statusLed")
        self.led.setProperty("connected", "false")
        h.addWidget(self.led)

        # 连接按钮
        self.btn = QPushButton("连  接")
        self.btn.setObjectName("btnConnect")
        self.btn.setProperty("connected", "false")
        self.btn.clicked.connect(self._on_click)
        h.addWidget(self.btn)

    def _on_click(self):
        if not self._connected:
            parity_map = {"None": "N", "Odd": "O", "Even": "E"}
            params = {
                "port":     self.cb_port.currentText(),
                "baudrate": int(self.cb_baud.currentText()),
                "bytesize": int(self.cb_data.currentText()),
                "stopbits": int(self.cb_stop.currentText()),
                "parity":   parity_map[self.cb_parity.currentText()],
                "slave_id": self.sp_slave.value(),
                "poll_ms":  self.sp_poll.value(),
            }
            self.connect_clicked.emit(params)
        else:
            self.disconnect_clicked.emit()

    def set_connected(self, ok: bool):
        self._connected = ok
        self.led.setProperty("connected", "true" if ok else "false")
        self.btn.setProperty("connected", "true" if ok else "false")
        self.btn.setText("断  开" if ok else "连  接")
        # 刷新样式
        self.led.style().unpolish(self.led)
        self.led.style().polish(self.led)
        self.btn.style().unpolish(self.btn)
        self.btn.style().polish(self.btn)

    @property
    def poll_ms(self):
        return self.sp_poll.value()


# ─────────────────────────────────────────────────────────────────
#  主窗口
# ─────────────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self, config_path: str):
        super().__init__()
        self.setWindowTitle("Modbus 上位机")
        self.resize(1280, 820)

        self.modbus = ModbusClient()
        self._poll_thread: Optional[QThread] = None
        self._poll_worker: Optional[PollingWorker] = None

        # 加载配置
        try:
            self.sheet_configs = load_config(config_path)
        except Exception as e:
            QMessageBox.critical(self, "配置文件错误", f"无法读取配置文件:\n{e}")
            sys.exit(1)

        self._build_ui()
        self.statusBar().showMessage(f"已加载配置: {os.path.basename(config_path)}"
                                     f"  |  共 {self._total_channels()} 个通道"
                                     f"  |  {len(self.sheet_configs)} 个 Tab")

    def _total_channels(self):
        return sum(len(v) for v in self.sheet_configs.values())

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        vbox = QVBoxLayout(central)
        vbox.setContentsMargins(0, 0, 0, 0)
        vbox.setSpacing(0)

        # 串口栏
        self.serial_bar = SerialBar()
        self.serial_bar.connect_clicked.connect(self._on_connect)
        self.serial_bar.disconnect_clicked.connect(self._on_disconnect)
        vbox.addWidget(self.serial_bar)

        # 分割线
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color: #2e333f;")
        vbox.addWidget(sep)

        # Tab 组
        self.tabs = QTabWidget()
        self.tab_pages: Dict[str, SheetTab] = {}
        vbox.addWidget(self.tabs)

        for sheet_name, configs in self.sheet_configs.items():
            page = SheetTab(configs)
            page.write_requested.connect(self._on_write)
            self.tabs.addTab(page, sheet_name)
            self.tab_pages[sheet_name] = page

        # 状态栏
        self.setStatusBar(QStatusBar())

    # ── Modbus 连接 ──────────────────────────────────────────────
    def _on_connect(self, params: dict):
        ok = self.modbus.connect(
            port=params["port"],
            baudrate=params["baudrate"],
            stopbits=params["stopbits"],
            bytesize=params["bytesize"],
            parity=params["parity"],
            slave_id=params["slave_id"],
        )
        self.serial_bar.set_connected(ok)
        if ok:
            self.statusBar().showMessage(
                f"已连接 {params['port']} @ {params['baudrate']} bps  "
                f"| 从站 {params['slave_id']}  | 轮询 {params['poll_ms']} ms")
            self._start_polling(params["poll_ms"], params["slave_id"])
        else:
            QMessageBox.warning(self, "连接失败", f"无法连接到 {params['port']}")

    def _on_disconnect(self):
        self._stop_polling()
        self.modbus.disconnect()
        self.serial_bar.set_connected(False)
        self.statusBar().showMessage("已断开连接")

    # ── 轮询线程 ─────────────────────────────────────────────────
    def _all_registers(self) -> List[int]:
        regs = []
        for configs in self.sheet_configs.values():
            for cfg in configs:
                regs.append(cfg.register)
        return list(set(regs))

    def _start_polling(self, interval_ms: int, slave_id: int):
        self._stop_polling()
        regs = self._all_registers()
        self._poll_worker = PollingWorker(self.modbus, regs, interval_ms)
        self._poll_thread = QThread()
        self._poll_worker.moveToThread(self._poll_thread)
        self._poll_thread.started.connect(self._poll_worker.run)
        self._poll_worker.values_ready.connect(self._on_values)
        self._poll_worker.finished.connect(self._poll_thread.quit)
        self._poll_thread.start()

    def _stop_polling(self):
        if self._poll_worker:
            self._poll_worker.stop()
        if self._poll_thread:
            self._poll_thread.quit()
            self._poll_thread.wait(2000)
        self._poll_worker = None
        self._poll_thread = None

    def _on_values(self, data: Dict[int, int]):
        for page in self.tab_pages.values():
            page.apply_values(data)

    # ── Modbus 写 ────────────────────────────────────────────────
    def _on_write(self, register: int, raw_value: int):
        ok = self.modbus.write_register(register, raw_value)
        msg = (f"写入成功: 寄存器 {register} = {raw_value}" if ok
               else f"写入失败: 寄存器 {register}")
        self.statusBar().showMessage(msg, 3000)

    def closeEvent(self, event):
        self._stop_polling()
        self.modbus.disconnect()
        event.accept()


# ─────────────────────────────────────────────────────────────────
#  入口
# ─────────────────────────────────────────────────────────────────
def main():
    # 配置文件路径：同目录下的 device-config.xlsx
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "device-config.xlsx")

    if not os.path.exists(config_path):
        print(f"[ERROR] 配置文件不存在: {config_path}")
        sys.exit(1)

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setStyleSheet(STYLE)

    win = MainWindow(config_path)
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
